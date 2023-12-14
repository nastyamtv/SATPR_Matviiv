import openpyxl
import numpy as np


def get_matrix_from_sheet(sheet):
    matrix = []

    for row in sheet.iter_rows(values_only=True):
        matrix.append(row)

    return np.array(matrix)


def calculate_priority_matrix(matrix):
    number_of_rows = matrix.shape[0]
    priorities_matrix = np.zeros((number_of_rows, 1))
    priorities_sum = 0

    for row in range(number_of_rows):
        priority = np.prod(matrix[row])
        priority = priority ** (1 / number_of_rows)
        priorities_sum += priority
        priorities_matrix[row, 0] = priority

    priorities_matrix /= priorities_sum
    return priorities_matrix


def create_criteria_priority_matrix(workbook):
    matrix = get_matrix_from_sheet(workbook.worksheets[0])
    matrix = calculate_priority_matrix(matrix)

    return matrix


def create_alternatives_priority_matrix(workbook):
    sheet_names = workbook.sheetnames
    number_of_sheets = len(sheet_names)

    tmp_matrix1 = get_matrix_from_sheet(workbook.worksheets[0])
    tmp_matrix2 = get_matrix_from_sheet(workbook.worksheets[1])

    number_of_columns = tmp_matrix1.shape[1]
    number_of_rows = tmp_matrix2.shape[0]

    alternative_priority_matrix = np.zeros((number_of_rows, number_of_columns))

    for i in range(1, number_of_sheets):
        matrix = get_matrix_from_sheet(workbook.worksheets[i])
        alternative_priority_matrix_column = calculate_priority_matrix(matrix)

        for j in range(number_of_rows):
            alternative_priority_matrix[j][i - 1] = alternative_priority_matrix_column[j][0]

    return alternative_priority_matrix


def display_results_matrix(matrix, text):
    for i, row in enumerate(matrix, start=1):
        for value in row:
            print(f"{text}{i}: {round(value, 4)}")


def find_and_display_result(matrix1, matrix2):
    results_matrix = np.dot(matrix1, matrix2)

    max_result_index = np.unravel_index(np.argmax(results_matrix), results_matrix.shape)

    result = round(results_matrix[max_result_index], 4)

    print('Priority of alternatives:')
    display_results_matrix(results_matrix, 'A')
    print()
    print(f'Max value {result} in alternative A{max_result_index[0] + 1}. '
          f'Then this is the best option.')


file_path = 'lab2.xlsx'

workbook = openpyxl.load_workbook(file_path)

criteria_priority_matrix = create_criteria_priority_matrix(workbook)
alternative_priority_matrix = create_alternatives_priority_matrix(workbook)

find_and_display_result(alternative_priority_matrix, criteria_priority_matrix)

workbook.close()
